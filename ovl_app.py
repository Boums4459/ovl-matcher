"""
OVL Matcher — Interface Web (multi-utilisateurs)
=================================================
Chaque upload reçoit un session_id UUID isolé.
Plusieurs collègues peuvent utiliser l'outil simultanément
sans interférence.

Usage local  : python ovl_app.py
Usage Render : gunicorn ovl_app:app
"""

import sys
import re
import os
import shutil
import json
import uuid
import threading
from pathlib import Path
from datetime import datetime

from flask import Flask, request, jsonify, send_file

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant. Installe-le : pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────

SHEET_OVL         = "OVL"
SHEET_SOURCE      = "Source_Word"
OVL_HEADER_MARKER = "Characteristic Name"
OVL_COL_CHAR_NAME = 1
OVL_COL_VALUE     = 4
OVL_COL_CHECK     = 6
SOURCE_COL_CODE   = 2
VAL_CONFIRMED     = "confirmed"
VAL_SELECT        = "select"
OVL_STOP_MARKERS  = {"Additional options to be added:"}
CODE_PATTERN      = re.compile(r'^[A-Z0-9][A-Z0-9_]{3,}$')
PORT              = 5050

# ─────────────────────────────────────────────────────────────
# MOTEUR DE TRAITEMENT
# ─────────────────────────────────────────────────────────────

class OVLProcessor:
    def __init__(self):
        self.logs = []

    def log(self, level, msg):
        self.logs.append({"level": level, "msg": msg})

    def extract_codes(self, ws):
        codes = set()
        for row in ws.iter_rows(values_only=True):
            raw = row[SOURCE_COL_CODE]
            if raw is None:
                continue
            code = str(raw).strip().upper()
            if code and CODE_PATTERN.match(code):
                codes.add(code)
        self.log("info", f"Source_Word : {len(codes)} codes extraits")
        return codes

    def find_data_rows(self, ws):
        header_row = None
        data_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if header_row is None:
                if row[OVL_COL_CHAR_NAME] == OVL_HEADER_MARKER:
                    header_row = i
                    self.log("info", f"OVL : header trouvé ligne {i}")
                continue
            if row[0] in OVL_STOP_MARKERS:
                break
            if all(cell is None for cell in row[:8]):
                break
            data_rows.append((i, row[OVL_COL_CHAR_NAME], row[OVL_COL_VALUE]))
        if header_row is None:
            raise ValueError(f"Header '{OVL_HEADER_MARKER}' introuvable dans '{SHEET_OVL}'")
        self.log("info", f"OVL : {len(data_rows)} lignes de données")
        return data_rows

    def compute_check(self, b_val, e_val, codes):
        def norm(v):
            if v is None:
                return None
            s = str(v).strip().upper()
            return s if s and s != "0" else None
        b, e = norm(b_val), norm(e_val)
        return VAL_CONFIRMED if (b and b in codes) or (e and e in codes) else VAL_SELECT

    def process(self, input_path: Path) -> Path:
        self.logs = []
        self.log("info", f"Fichier reçu : {input_path.name}")

        # Backup dans le même dossier de session
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = input_path.parent / f"{input_path.stem}_backup_{ts}{input_path.suffix}"
        shutil.copy2(input_path, backup)
        self.log("info", f"Backup : {backup.name}")

        # Étape 1 — Source_Word
        self.log("step", "Étape 1/3 — Extraction des codes Source_Word")
        wb_ro = load_workbook(str(input_path), read_only=True, data_only=True)
        if SHEET_SOURCE not in wb_ro.sheetnames:
            raise ValueError(f"Feuille '{SHEET_SOURCE}' introuvable. Feuilles : {wb_ro.sheetnames}")
        codes = self.extract_codes(wb_ro[SHEET_SOURCE])
        wb_ro.close()
        if not codes:
            raise ValueError(f"Aucun code extrait de '{SHEET_SOURCE}'")

        # Étape 2 — Lecture OVL
        self.log("step", "Étape 2/3 — Lecture du tableau OVL")
        wb_ro2 = load_workbook(str(input_path), read_only=True, data_only=True)
        if SHEET_OVL not in wb_ro2.sheetnames:
            raise ValueError(f"Feuille '{SHEET_OVL}' introuvable. Feuilles : {wb_ro2.sheetnames}")
        data_rows = self.find_data_rows(wb_ro2[SHEET_OVL])
        wb_ro2.close()

        # Étape 3 — Écriture
        self.log("step", "Étape 3/3 — Écriture colonne Check")
        wb_rw = load_workbook(str(input_path))
        ws = wb_rw[SHEET_OVL]
        confirmed, select = 0, 0
        for row_num, b_val, e_val in data_rows:
            val = self.compute_check(b_val, e_val, codes)
            ws[f"G{row_num}"] = val
            if val == VAL_CONFIRMED:
                confirmed += 1
            else:
                select += 1

        output = input_path.parent / f"{input_path.stem}_OVL_updated{input_path.suffix}"
        wb_rw.save(str(output))
        wb_rw.close()

        self.log("success", f"✅ confirmed : {confirmed}")
        self.log("success", f"⬜ select    : {select}")
        self.log("success", f"Fichier prêt : {output.name}")
        return output


# ─────────────────────────────────────────────────────────────
# FLASK APP
# ─────────────────────────────────────────────────────────────

app = Flask(__name__)
UPLOAD_DIR = Path(__file__).parent / "uploads_tmp"
UPLOAD_DIR.mkdir(exist_ok=True)

# Stockage des fichiers produits par session (en mémoire)
# Chaque utilisateur reçoit un UUID isolé → pas de collision
session_outputs = {}


@app.route("/")
def index():
    return HTML_PAGE


@app.route("/process", methods=["POST"])
def process_file():
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier reçu"}), 400

    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Format non supporté. Dépose un .xlsx"}), 400

    # Dossier isolé par session
    session_id = str(uuid.uuid4())
    session_dir = UPLOAD_DIR / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    save_path = session_dir / f.filename
    f.save(str(save_path))

    proc = OVLProcessor()
    try:
        output = proc.process(save_path)
        session_outputs[session_id] = output
        return jsonify({
            "logs": proc.logs,
            "output": output.name,
            "session_id": session_id
        })
    except Exception as e:
        proc.log("error", str(e))
        return jsonify({"logs": proc.logs, "error": str(e)}), 400


@app.route("/download/<session_id>")
def download(session_id):
    output = session_outputs.get(session_id)
    if output and output.exists():
        return send_file(str(output), as_attachment=True)
    return "Fichier introuvable ou session expirée", 404


# ─────────────────────────────────────────────────────────────
# HTML (interface embarquée)
# ─────────────────────────────────────────────────────────────

HTML_PAGE = """<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>OVL Matcher</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

  * { margin: 0; padding: 0; box-sizing: border-box; }

  :root {
    --bg: #0f1117;
    --surface: #1a1d27;
    --border: #2a2e3b;
    --text: #e4e6ed;
    --text-dim: #8b8fa3;
    --accent: #3b82f6;
    --accent-glow: rgba(59, 130, 246, 0.15);
    --green: #22c55e;
    --green-dim: rgba(34, 197, 94, 0.12);
    --red: #ef4444;
    --radius: 10px;
  }

  body {
    font-family: 'Inter', -apple-system, sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    align-items: center;
    padding: 48px 24px;
  }

  .container { max-width: 640px; width: 100%; }

  h1 { font-size: 1.5rem; font-weight: 700; letter-spacing: -0.02em; margin-bottom: 4px; }
  .subtitle { color: var(--text-dim); font-size: 0.85rem; margin-bottom: 36px; }

  .drop-zone {
    border: 2px dashed var(--border);
    border-radius: var(--radius);
    padding: 48px 24px;
    text-align: center;
    cursor: pointer;
    transition: all 0.2s;
    background: var(--surface);
    margin-bottom: 16px;
  }
  .drop-zone:hover, .drop-zone.drag-over { border-color: var(--accent); background: var(--accent-glow); }
  .drop-zone.has-file { border-color: var(--green); border-style: solid; background: var(--green-dim); }
  .drop-zone p { color: var(--text-dim); font-size: 0.9rem; }
  .drop-zone .filename { color: var(--green); font-weight: 600; font-size: 0.95rem; }
  .drop-zone .icon { font-size: 2rem; margin-bottom: 12px; }

  input[type="file"] { display: none; }

  .btn {
    width: 100%;
    padding: 14px;
    border: none;
    border-radius: var(--radius);
    font-family: inherit;
    font-size: 0.95rem;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.15s;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
  }
  .btn-primary { background: var(--accent); color: white; }
  .btn-primary:hover:not(:disabled) { background: #2563eb; }
  .btn-primary:disabled { opacity: 0.35; cursor: not-allowed; }
  .btn-download {
    background: var(--green);
    color: #000;
    margin-top: 12px;
    text-decoration: none;
    display: none;
  }
  .btn-download:hover { background: #16a34a; }
  .btn-download.visible { display: flex; }

  .log-panel {
    margin-top: 24px;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    display: none;
  }
  .log-panel.visible { display: block; }
  .log-header {
    padding: 12px 16px;
    font-size: 0.8rem;
    font-weight: 600;
    color: var(--text-dim);
    text-transform: uppercase;
    letter-spacing: 0.05em;
    border-bottom: 1px solid var(--border);
  }
  .log-body {
    padding: 16px;
    max-height: 400px;
    overflow-y: auto;
    font-family: 'SF Mono', 'Cascadia Code', monospace;
    font-size: 0.82rem;
    line-height: 1.7;
  }
  .log-line.info    { color: var(--text-dim); }
  .log-line.step    { color: var(--accent); font-weight: 500; }
  .log-line.success { color: var(--green); }
  .log-line.error   { color: var(--red); }

  .spinner {
    display: inline-block; width: 16px; height: 16px;
    border: 2px solid rgba(255,255,255,0.3);
    border-top-color: white; border-radius: 50%;
    animation: spin 0.6s linear infinite;
  }
  @keyframes spin { to { transform: rotate(360deg); } }
</style>
</head>
<body>
<div class="container">
  <h1>OVL Matcher</h1>
  <p class="subtitle">Dépose le fichier Excel (avec la feuille Source_Word remplie)</p>

  <div class="drop-zone" id="dropZone">
    <div class="icon">📄</div>
    <p>Glisse le .xlsx ici ou clique pour parcourir</p>
  </div>
  <input type="file" id="fileInput" accept=".xlsx,.xlsm">

  <button class="btn btn-primary" id="btnRun" disabled>Lancer le traitement</button>
  <a class="btn btn-download" id="btnDownload" href="#">Télécharger le fichier traité</a>

  <div class="log-panel" id="logPanel">
    <div class="log-header">Journal</div>
    <div class="log-body" id="logBody"></div>
  </div>
</div>

<script>
  const dropZone    = document.getElementById('dropZone');
  const fileInput   = document.getElementById('fileInput');
  const btnRun      = document.getElementById('btnRun');
  const btnDownload = document.getElementById('btnDownload');
  const logPanel    = document.getElementById('logPanel');
  const logBody     = document.getElementById('logBody');

  let selectedFile = null;

  dropZone.addEventListener('click', () => fileInput.click());
  dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('drag-over'); });
  dropZone.addEventListener('dragleave', () => dropZone.classList.remove('drag-over'));
  dropZone.addEventListener('drop', e => {
    e.preventDefault();
    dropZone.classList.remove('drag-over');
    if (e.dataTransfer.files.length) setFile(e.dataTransfer.files[0]);
  });
  fileInput.addEventListener('change', () => { if (fileInput.files.length) setFile(fileInput.files[0]); });

  function setFile(f) {
    if (!f.name.match(/\\.xlsx?$/i)) { alert('Fichier .xlsx uniquement'); return; }
    selectedFile = f;
    dropZone.classList.add('has-file');
    dropZone.innerHTML = '<div class="icon">✅</div><p class="filename">' + f.name + '</p>'
      + '<p style="margin-top:6px;color:var(--text-dim);font-size:0.8rem">'
      + (f.size/1024).toFixed(0) + ' Ko — cliquer pour changer</p>';
    btnRun.disabled = false;
    btnDownload.classList.remove('visible');
    logPanel.classList.remove('visible');
  }

  btnRun.addEventListener('click', async () => {
    if (!selectedFile) return;
    btnRun.disabled = true;
    btnRun.innerHTML = '<span class="spinner"></span> Traitement en cours…';
    btnDownload.classList.remove('visible');
    logBody.innerHTML = '';
    logPanel.classList.add('visible');
    addLog('info', 'Envoi du fichier au serveur…');

    const fd = new FormData();
    fd.append('file', selectedFile);

    try {
      const resp = await fetch('/process', { method: 'POST', body: fd });
      const data = await resp.json();

      if (data.logs) data.logs.forEach(l => addLog(l.level, l.msg));

      if (data.error) {
        addLog('error', 'Erreur : ' + data.error);
      } else {
        addLog('success', '— Terminé —');
        // Lien de téléchargement isolé par session_id
        btnDownload.href = '/download/' + data.session_id;
        btnDownload.classList.add('visible');
      }
    } catch (err) {
      addLog('error', 'Erreur réseau : ' + err.message);
    }

    btnRun.disabled = false;
    btnRun.innerHTML = 'Lancer le traitement';
  });

  function addLog(level, msg) {
    const div = document.createElement('div');
    div.className = 'log-line ' + level;
    div.textContent = msg;
    logBody.appendChild(div);
    logBody.scrollTop = logBody.scrollHeight;
  }
</script>
</body>
</html>
"""

# ─────────────────────────────────────────────────────────────
# LANCEMENT
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", PORT))
    print(f"OVL Matcher — http://localhost:{port}")
    print("Ctrl+C pour arrêter\n")
    app.run(host="0.0.0.0", port=port, debug=False)
